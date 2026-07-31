"""Getting data back out of OrgView.

Until now the only file this app could produce was a blank import template, so a
corrected census had nowhere to go. These exports close that loop — most
importantly the census one, which round-trips: its columns are the canonical
import field names, so a corrected file can be handed straight back to the
portfolio company and re-uploaded here without remapping.

Two rules every exporter follows, because a spreadsheet is the easiest way to
leak something:

- **Pay columns are omitted, not blanked**, for anyone without pay access. A
  blank column still tells you the field exists and invites a follow-up.
- **Branch-restricted users export their branch only**, matching what the chart
  and the API already give them.
"""
from __future__ import annotations

import csv
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse

from ..models import Employee, ScenarioPosition, StructureCorrection  # noqa: F401
from ..parsers import STANDARD_FIELDS
from .costing import cost_of
from .tree_builder import build_tree

#: Import field name -> Employee attribute, where they differ.
_FIELD_SOURCE = {"supervisor_id": "raw_supervisor_id"}

#: The two columns that reveal pay. `revenue_attribution` is commercially
#: sensitive in the same way, so it travels with them.
PAY_COLUMNS = ("annual_salary", "fully_loaded_cost", "revenue_attribution")


def _fmt(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _safe(name: str) -> str:
    keep = "-_. "
    cleaned = "".join(c for c in str(name) if c.isalnum() or c in keep).strip()
    return (cleaned or "export").replace(" ", "-").lower()


# ---------------------------------------------------------------------------
# The corrected census
# ---------------------------------------------------------------------------

def census_sheet(snapshot, *, include_pay=True, branch_root=None,
                 include_excluded=False) -> tuple[str, list[str], list[list]]:
    """The snapshot as it stands *after* corrections, ready to re-import.

    Adds one non-standard trailing column, ``_corrections``, naming what this app
    changed about each row. It starts with an underscore so a re-import ignores
    it, but it is the whole point of handing the file back: the recipient can see
    exactly which rows we disagreed with their export about.
    """
    columns = [f for f, _, _ in STANDARD_FIELDS]
    if not include_pay:
        columns = [c for c in columns if c not in PAY_COLUMNS]

    qs = Employee.objects.filter(snapshot=snapshot)
    if not include_excluded:
        qs = qs.exclude(employee_status=Employee.EXCLUDED_STATUS)
    if branch_root:
        allowed = _branch_ids(snapshot, branch_root)
        qs = qs.filter(employee_id__in=allowed)

    corrections = _corrections_by_employee(snapshot.company)

    rows = []
    for emp in qs.order_by("last_name", "first_name"):
        row = [_fmt(getattr(emp, _FIELD_SOURCE.get(c, c), "")) for c in columns]
        row.append(", ".join(corrections.get(emp.employee_id, [])))
        rows.append(row)

    return "Census", columns + ["_corrections"], rows


def _corrections_by_employee(company) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for c in StructureCorrection.objects.filter(company=company, is_active=True):
        out.setdefault(c.employee_id, []).append(c.get_kind_display())
    return out


def _branch_ids(snapshot, branch_root) -> set[str]:
    tree = build_tree(snapshot.id, root_employee_id=branch_root)
    ids, stack = set(), (list(tree) if isinstance(tree, list) else [tree])
    while stack:
        node = stack.pop()
        if not node:
            continue
        ids.add(node["employee_id"])
        stack.extend(node.get("children") or [])
    return ids


# ---------------------------------------------------------------------------
# The corrections ledger
# ---------------------------------------------------------------------------

def corrections_sheet(company, snapshot) -> tuple[str, list[str], list[list]]:
    names = {}
    if snapshot:
        names = {
            e["employee_id"]: f'{e["first_name"]} {e["last_name"]}'.strip()
            for e in Employee.objects.filter(snapshot=snapshot)
            .values("employee_id", "first_name", "last_name")
        }

    headers = ["employee_id", "name", "correction", "status", "detail",
               "before", "after", "note", "active", "by", "updated"]
    rows = []
    for c in StructureCorrection.objects.filter(company=company).select_related("created_by"):
        rows.append([
            c.employee_id,
            names.get(c.employee_id, ""),
            c.get_kind_display(),
            c.get_replay_status_display(),
            c.replay_detail,
            _kv(c.before),
            _kv(c.after),
            c.note,
            "Yes" if c.is_active else "No",
            c.created_by.get_username() if c.created_by else "",
            c.updated_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return "Corrections", headers, rows


# ---------------------------------------------------------------------------
# The HR action list
# ---------------------------------------------------------------------------

_FIELD_LABELS = {f: label for f, label, _ in STANDARD_FIELDS}
_FIELD_LABELS.update({
    "raw_supervisor_id": "Supervisor",
    "supervisor_id": "Supervisor",
    "employee_status": "Employee Status",
})

#: What each replay status means for whoever has to key the change in.
_ACTION_STATUS = {
    StructureCorrection.ReplayStatus.APPLIED:
        ("Action needed", "Not yet reflected in the census export."),
    StructureCorrection.ReplayStatus.DRIFTED:
        ("Check first", "The export has changed since this was recorded — it may "
                        "already be done, or done differently."),
    StructureCorrection.ReplayStatus.CONFLICT:
        ("Blocked", "Could not be applied to the latest census."),
    StructureCorrection.ReplayStatus.STALE:
        ("No longer applies", "This person is not in the latest census."),
    StructureCorrection.ReplayStatus.RESOLVED:
        ("Already done", "The export now matches — no action needed."),
}


def hr_actions_sheet(company, snapshot, *, include_inactive=False) -> tuple[str, list[str], list[list]]:
    """A worklist for whoever has to key these changes into the HR system.

    One row per *individual change*, not per correction — an attribute correction
    touching three fields is three rows to tick off, and eliminating a manager
    produces a manager change for every one of their reports, which is real work
    for HR that nothing else in this app would have surfaced.

    ``before`` is what the payroll export said and ``after`` is what we corrected
    it to, so the two map directly onto "your system says X, it should be Y".
    """
    names = {}
    if snapshot:
        names = {
            e["employee_id"]: f'{e["first_name"]} {e["last_name"]}'.strip()
            for e in Employee.objects.filter(snapshot=snapshot)
            .values("employee_id", "first_name", "last_name")
        }

    def who(eid):
        if not eid:
            return ""
        name = names.get(str(eid))
        return f"{name} ({eid})" if name else str(eid)

    headers = ["priority", "employee_id", "name", "action", "field",
               "current value in HR system", "change to", "why", "status",
               "recorded by", "recorded"]
    rows = []

    qs = StructureCorrection.objects.filter(company=company).select_related("created_by")
    if not include_inactive:
        qs = qs.filter(is_active=True)

    K = StructureCorrection.Kind
    for c in qs:
        status, status_note = _ACTION_STATUS.get(
            c.replay_status, ("Review", c.replay_detail or ""))
        # Anything already done, or that can't be done, sinks below live work.
        priority = 1 if status == "Action needed" else 2 if status == "Check first" else 3
        before, after = c.before or {}, c.after or {}
        stamp = c.updated_at.strftime("%Y-%m-%d")
        by = c.created_by.get_username() if c.created_by else ""
        note = c.note or ""

        def emit(eid, action, field, frm, to, why=None):
            rows.append([
                priority, eid, names.get(eid, ""), action, field, frm, to,
                why if why is not None else note,
                f"{status}{' — ' + status_note if status_note else ''}", by, stamp,
            ])

        if c.kind == K.REPARENT:
            emit(c.employee_id, "Change manager", "Supervisor",
                 who(before.get("raw_supervisor_id")), who(after.get("raw_supervisor_id")))

        elif c.kind == K.SET_ROOT:
            emit(c.employee_id, "Change manager", "Supervisor",
                 who(before.get("raw_supervisor_id")), "(none — top of organisation)")

        elif c.kind == K.ATTRIBUTE:
            for field, value in after.items():
                emit(c.employee_id, "Correct field",
                     _FIELD_LABELS.get(field, field),
                     _fmt(before.get(field, "")), _fmt(value))

        elif c.kind == K.ADD_PERSON:
            label = " ".join(
                str(after.get(f, "")) for f in ("first_name", "last_name")).strip()
            emit(c.employee_id, "Add to HR system", "Record",
                 "(missing from export)",
                 f"{label or after.get('job_title') or c.employee_id}"
                 f" — reports to {who(after.get('raw_supervisor_id'))}")

        elif c.kind in (K.EXCLUDE, K.ELIMINATE):
            eliminating = c.kind == K.ELIMINATE
            emit(c.employee_id,
                 "Remove position" if eliminating else "Remove duplicate/ghost record",
                 "Record",
                 "(on the org chart)",
                 "(remove — role eliminated)" if eliminating
                 else "(remove — duplicate or ghost row in the export)")
            # Everyone the removal re-homed is a manager change HR has to make too.
            reassign = after.get("reassign") or {}
            fallback = before.get("raw_supervisor_id")
            for child, prior in (before.get("reports") or {}).items():
                emit(child, "Change manager", "Supervisor", who(prior),
                     who(reassign.get(child) or fallback),
                     f"{names.get(c.employee_id, c.employee_id)}'s position was "
                     f"{'eliminated' if eliminating else 'removed'}"
                     + (f" — {note}" if note else ""))

    # Live work first, then by person, so it reads as a to-do list.
    rows.sort(key=lambda r: (r[0], r[2] or "", r[3]))
    return "HR actions", headers, rows


def _kv(blob) -> str:
    if not isinstance(blob, dict) or not blob:
        return ""
    parts = []
    for key, value in blob.items():
        if isinstance(value, dict):
            value = "; ".join(f"{k}->{v}" for k, v in value.items()) or "—"
        parts.append(f"{key}={value}")
    return ", ".join(parts)


# ---------------------------------------------------------------------------
# A scenario
# ---------------------------------------------------------------------------

def scenario_sheets(scenario, summary, *, include_pay=True) -> list[tuple]:
    """Impact summary, change log and full roster — one sheet each."""
    base, scen, deltas = summary["baseline"], summary["scenario"], summary["deltas"]

    impact = [
        ["Scenario", scenario.name],
        ["Description", scenario.description],
        ["Baseline census", str(scenario.base_snapshot) if scenario.base_snapshot else "—"],
        ["Exported", scenario.updated_at.strftime("%Y-%m-%d %H:%M")],
        [],
        ["Metric", "Baseline", "Scenario", "Change"],
        ["Headcount", base["headcount"], scen["headcount"], deltas["headcount"]],
        ["Layers", base["layers"], scen["layers"], deltas["layers"]],
        ["Avg span of control", base["avg_span"], scen["avg_span"], deltas["avg_span"]],
    ]
    if include_pay:
        impact += [
            ["Total loaded cost", _fmt(base["total_cost"]), _fmt(scen["total_cost"]),
             _fmt(deltas["total_cost"])],
            [],
            ["Investment", _fmt(summary["totals"]["investment"])],
            ["Savings", _fmt(summary["totals"]["savings"])],
            ["Net annual impact", _fmt(summary["totals"]["net"])],
        ]

    ledger_headers = ["change", "employee_id", "name", "job_title", "department", "vacant"]
    if include_pay:
        ledger_headers.append("cost_impact")
    ledger_headers.append("note")
    ledger_rows = []
    for item in summary["ledger"]:
        row = [item["change_label"], item["employee_id"], item["name"],
               item["job_title"], item["department"],
               "Yes" if item["is_vacant"] else "No"]
        if include_pay:
            row.append(_fmt(item["cost_impact"]))
        row.append(item["note"])
        ledger_rows.append(row)

    roster_cols = ["employee_id", "first_name", "last_name", "raw_supervisor_id",
                   "job_title", "management_level", "department", "site_location",
                   "city", "state", "entity", "employee_status", "employee_type",
                   "pay_type"]
    if include_pay:
        roster_cols += ["annual_salary", "fully_loaded_cost"]
    roster_cols += ["change_type", "is_vacant", "prior_supervisor_id", "note"]

    roster_rows = []
    ct = ScenarioPosition.ChangeType
    for p in scenario.positions.exclude(change_type=ct.REMOVED).order_by("last_name", "first_name"):
        roster_rows.append([_fmt(getattr(p, c, "")) for c in roster_cols])

    return [
        ("Impact", None, impact),
        ("Change log", ledger_headers, ledger_rows),
        ("Positions", roster_cols, roster_rows),
    ]


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def csv_response(filename: str, headers, rows) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_safe(filename)}.csv"'
    # Excel needs the BOM to read UTF-8 accented names correctly.
    response.write("﻿")
    writer = csv.writer(response)
    if headers:
        writer.writerow(headers)
    writer.writerows(rows)
    return response


def xlsx_response(filename: str, sheets) -> HttpResponse:
    """sheets: [(title, headers|None, rows)] — a None header row is a free-form sheet."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=str(title)[:31])
        widths: dict[int, int] = {}

        def put(values, bold=False):
            ws.append(list(values))
            for i, value in enumerate(values, start=1):
                if bold:
                    ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
                widths[i] = min(48, max(widths.get(i, 10), len(str(value or "")) + 2))

        if headers:
            put(headers, bold=True)
            ws.freeze_panes = "A2"
        for row in rows:
            put(row)

        for i, width in widths.items():
            ws.column_dimensions[get_column_letter(i)].width = width
        if headers:
            for cell in ws[1]:
                cell.alignment = Alignment(vertical="center")

    buf = BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_safe(filename)}.xlsx"'
    return response
