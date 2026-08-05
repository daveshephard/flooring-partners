import csv
import io
import json
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse

from accounts.models import AppDefinition, CompanyProfile, UserProfile

from .models import (
    AppPermission, CensusSnapshot, ChartGroup, Company, CorrectionReplayLog, Employee, Scenario,
    ScenarioPosition, StructureCorrection,
)
from .services import changeset as CS
from .services import corrections as C
from .services import scenarios as S
from .services.company_sync import ensure_org_view_company, sync_from_accounts
from .services.costing import cost_of
from .services.tree_builder import (
    _build_lookups, build_tree, build_tree_from_rows, detect_cycles, strip_aggregation,
)


class CompanySyncTests(TestCase):
    def test_ensure_creates_missing_company(self):
        company, created = ensure_org_view_company("Flooring Partners")
        self.assertTrue(created)
        self.assertEqual(company.name, "Flooring Partners")
        self.assertEqual(company.slug, "flooring-partners")
        self.assertTrue(company.is_active)

    def test_ensure_is_idempotent(self):
        ensure_org_view_company("Flooring Partners")
        company, created = ensure_org_view_company("Flooring Partners")
        self.assertFalse(created)
        self.assertEqual(Company.objects.filter(slug="flooring-partners").count(), 1)

    def test_ensure_ignores_blank(self):
        company, created = ensure_org_view_company("   ")
        self.assertIsNone(company)
        self.assertFalse(created)

    def test_sync_from_accounts_backfills_active_only(self):
        CompanyProfile.objects.create(name="Flooring Partners", is_active=True)
        CompanyProfile.objects.create(name="Dormant Co", is_active=False)
        # Simulate the pre-signal state (rows added before the sync existed).
        Company.objects.all().delete()
        created, existing = sync_from_accounts()
        self.assertIn("Flooring Partners", created)
        self.assertNotIn("Dormant Co", created)
        self.assertTrue(Company.objects.filter(slug="flooring-partners").exists())
        self.assertFalse(Company.objects.filter(slug="dormant-co").exists())

    def test_sync_dry_run_writes_nothing(self):
        CompanyProfile.objects.create(name="Flooring Partners", is_active=True)
        Company.objects.all().delete()  # pre-signal state
        created, existing = sync_from_accounts(dry_run=True)
        self.assertIn("Flooring Partners", created)
        self.assertFalse(Company.objects.filter(slug="flooring-partners").exists())

    def test_signal_creates_company_on_profile_save(self):
        CompanyProfile.objects.create(name="Signal Co", is_active=True)
        self.assertTrue(Company.objects.filter(slug="signal-co").exists())


class ScenarioEngineTests(TestCase):
    """Cloning, structural edits, and the cost-impact summary."""

    def setUp(self):
        self.company = Company.objects.create(name="Flooring Partners", slug="flooring-partners")
        self.snap = CensusSnapshot.objects.create(
            company=self.company, original_filename="census.csv",
            status=CensusSnapshot.Status.ACTIVE, is_current=True, employee_count=4,
        )
        # CEO -> VP -> {IC1, IC2}
        rows = [
            ("E1", "Ada",  "Root",  None, "CEO",        "300000"),
            ("E2", "Bo",   "Vice",  "E1", "VP Ops",     "200000"),
            ("E3", "Cy",   "One",   "E2", "Technician", "100000"),
            ("E4", "Di",   "Two",   "E2", "Technician", "110000"),
        ]
        for eid, fn, ln, sup, title, salary in rows:
            Employee.objects.create(
                snapshot=self.snap, employee_id=eid, first_name=fn, last_name=ln,
                raw_supervisor_id=sup, job_title=title, annual_salary=Decimal(salary),
            )

    def _new_scenario(self):
        return S.create_scenario(company=self.company, base_snapshot=self.snap, name="Reorg")

    def test_create_clones_all_positions_with_baseline_cost(self):
        sc = self._new_scenario()
        self.assertEqual(sc.positions.count(), 4)
        vp = sc.positions.get(employee_id="E2")
        self.assertEqual(vp.source_employee_id, "E2")
        self.assertEqual(vp.change_type, ScenarioPosition.ChangeType.UNCHANGED)
        self.assertEqual(vp.baseline_cost, Decimal("200000"))

    def test_baseline_summary_matches_snapshot(self):
        sc = self._new_scenario()
        summ = S.scenario_summary(sc)
        self.assertEqual(summ["baseline"]["headcount"], 4)
        self.assertEqual(summ["baseline"]["total_cost"], Decimal("710000"))
        self.assertEqual(summ["baseline"]["layers"], 3)

    def test_eliminate_reparents_reports_and_books_savings(self):
        sc = self._new_scenario()
        S.eliminate_position(sc.positions.get(employee_id="E2"))
        # E3/E4 now report to E1 (the VP's manager)
        self.assertEqual(sc.positions.get(employee_id="E3").raw_supervisor_id, "E1")
        self.assertEqual(sc.positions.get(employee_id="E4").raw_supervisor_id, "E1")
        summ = S.scenario_summary(sc)
        self.assertEqual(summ["scenario"]["headcount"], 3)
        self.assertEqual(summ["totals"]["savings"], Decimal("200000"))
        self.assertEqual(summ["totals"]["net"], Decimal("-200000"))

    def test_add_position_books_investment(self):
        sc = self._new_scenario()
        S.add_position(sc, supervisor_id="E1", job_title="VP Finance",
                       annual_salary=Decimal("150000"), is_vacant=True)
        summ = S.scenario_summary(sc)
        self.assertEqual(summ["scenario"]["headcount"], 5)
        self.assertEqual(summ["totals"]["investment"], Decimal("150000"))
        self.assertEqual(summ["totals"]["net"], Decimal("150000"))

    def test_edit_pay_books_delta(self):
        sc = self._new_scenario()
        S.edit_position(sc.positions.get(employee_id="E3"), {"annual_salary": Decimal("120000")})
        summ = S.scenario_summary(sc)
        modified = [x for x in summ["ledger"] if x["change_type"] == ScenarioPosition.ChangeType.MODIFIED]
        self.assertEqual(len(modified), 1)
        self.assertEqual(modified[0]["cost_impact"], Decimal("20000"))

    def test_reassign_cycle_guard(self):
        sc = self._new_scenario()
        # E1 (CEO) cannot report to E3, which is in E1's subtree.
        with self.assertRaises(ValueError):
            S.reassign_manager(sc.positions.get(employee_id="E1"), "E3")

    def test_scenario_tree_excludes_eliminated(self):
        sc = self._new_scenario()
        S.eliminate_position(sc.positions.get(employee_id="E4"))
        tree = S.build_scenario_tree(sc)
        root = tree if isinstance(tree, dict) else tree[0]
        self.assertEqual(root["metrics"]["headcount"], 3)

    def test_added_then_eliminated_disappears(self):
        sc = self._new_scenario()
        pos = S.add_position(sc, supervisor_id="E1", job_title="Temp")
        result = S.eliminate_position(pos)
        self.assertIsNone(result)
        self.assertEqual(sc.positions.count(), 4)


# ===========================================================================
# Shared fixtures for the editing work
# ===========================================================================

class OrgFixtureMixin:
    """A small, deliberately messy census.

    E1 (CEO)
      ├─ E2 (VP Ops) ─ E3, E4
      └─ E5 (CFO)    ─ E6
    X1 points at a supervisor who isn't in the file — the commonest defect, and
    the reason X1 is invisible on today's chart.
    """

    ROWS = [
        # eid,  first, last,    supervisor, title,        salary,   loaded
        ("E1", "Ada",  "Root",   None,   "CEO",          "300000", "380000"),
        ("E2", "Bo",   "Vice",   "E1",   "VP Ops",       "200000", None),
        ("E3", "Cy",   "One",    "E2",   "Technician",   "100000", "128000"),
        ("E4", "Di",   "Two",    "E2",   "Technician",   "110000", None),
        ("E5", "El",   "Three",  "E1",   "CFO",          "250000", "310000"),
        ("E6", "Fi",   "Four",   "E5",   "Accountant",   "85000",  None),
        ("X1", "Zed",  "Lost",   "9999", "Route Super",  "95000",  None),
    ]

    def make_company(self, name="Flooring Partners", slug="flooring-partners"):
        return Company.objects.create(name=name, slug=slug)

    def make_snapshot(self, company, rows=None, label="Q1", current=True):
        snap = CensusSnapshot.objects.create(
            company=company, label=label, original_filename="census.csv",
            status=CensusSnapshot.Status.ACTIVE, is_current=current,
        )
        for eid, fn, ln, sup, title, salary, loaded in (rows if rows is not None else self.ROWS):
            Employee.objects.create(
                snapshot=snap, employee_id=eid, first_name=fn, last_name=ln,
                raw_supervisor_id=sup, job_title=title,
                annual_salary=Decimal(salary) if salary else None,
                fully_loaded_cost=Decimal(loaded) if loaded else None,
                raw_data={"employee_id": eid, "supervisor_id": sup or ""},
            )
        snap.employee_count = snap.employees.count()
        snap.save(update_fields=["employee_count"])
        C.resolve_supervisor_fks(snap)
        return snap

    def make_user(self, username, company, role="admin", branch=None, staff=False):
        User = get_user_model()
        user = User.objects.create_user(username, password="pw", is_staff=staff)
        profile_company, _ = CompanyProfile.objects.get_or_create(name="Test Co")
        profile = UserProfile.objects.create(user=user, company=profile_company)
        app = AppDefinition.objects.filter(slug="org-view").first()
        if app:
            profile.assigned_apps.add(app)
        if not staff:
            AppPermission.objects.create(
                user=user, company=company, role=role, branch_root_employee_id=branch,
            )
        return user

    def correction(self, company, employee_id, kind, before, after, **kw):
        return StructureCorrection.objects.create(
            company=company, employee_id=employee_id, kind=kind,
            before=before, after=after, **kw,
        )


def _flatten_ids(tree):
    ids = set()
    stack = list(tree) if isinstance(tree, list) else [tree]
    while stack:
        n = stack.pop()
        ids.add(n["employee_id"])
        stack.extend(n.get("children") or [])
    return ids


# ===========================================================================
# Phase 6 §1 — the load-bearing tree_builder fixes
# ===========================================================================

class TreeBuilderFixTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()

    def test_cost_definition_is_shared(self):
        """tree_builder and scenarios must agree where the two columns differ.

        Before this fix the chart card summed annual_salary while the scenario
        impact panel summed fully_loaded_cost, so the two could never reconcile.
        """
        snap = self.make_snapshot(self.company)
        tree = strip_aggregation(build_tree(snap.id))
        scenario = S.create_scenario(company=self.company, base_snapshot=snap, name="X")
        summary = S.scenario_summary(scenario)

        # X1 is unreachable, so the chart's root rollup excludes it; compare like
        # for like by subtracting it from the scenario baseline.
        x1_cost = cost_of({"fully_loaded_cost": None, "annual_salary": Decimal("95000")})
        self.assertEqual(
            Decimal(str(tree["metrics"]["total_labor_cost"])),
            summary["baseline"]["total_cost"] - x1_cost,
        )
        # And the rollup must use loaded cost where present, not salary.
        self.assertEqual(tree["metrics"]["total_labor_cost"], 1213000.0)

    def test_build_tree_survives_cycle(self):
        """A supervisor loop used to recurse forever — a hung worker, not an error."""
        rows = [
            ("A", "A", "One",   "C",  "Mgr", "100", None),
            ("B", "B", "Two",   "A",  "Mgr", "100", None),
            ("C", "C", "Three", "B",  "Mgr", "100", None),
            ("R", "R", "Root",  None, "CEO", "100", None),
        ]
        snap = self.make_snapshot(self.company, rows=rows)
        report = {}
        tree = build_tree(snap.id, report=report)   # must return, not hang
        self.assertIsNotNone(tree)
        self.assertEqual(len(report["cycles"]), 1)
        self.assertEqual(set(report["cycles"][0]), {"A", "B", "C"})
        for eid in ("A", "B", "C"):
            self.assertEqual(report["unreachable"][eid], "in_cycle")

    def test_unreachable_rows_reported_with_reasons(self):
        rows = [
            ("R",  "R", "Root",   None,   "CEO", "100", None),
            ("K",  "K", "Kid",    "R",    "IC",  "100", None),
            ("O1", "O", "Orphan", "NOPE", "IC",  "100", None),   # supervisor_not_found
            ("O2", "O", "Second", None,   "IC",  "100", None),   # extra_root
            ("O3", "O", "Selfy",  "O3",   "IC",  "100", None),   # self_referential
            ("C1", "C", "Loop1",  "C2",   "IC",  "100", None),   # in_cycle
            ("C2", "C", "Loop2",  "C1",   "IC",  "100", None),
        ]
        emp_rows = [{
            "employee_id": eid, "first_name": fn, "last_name": ln,
            "raw_supervisor_id": sup, "job_title": t, "management_level": "",
            "department": "", "entity": "", "city": "", "state": "",
            "site_location": "", "employee_type": "", "pay_type": "",
            "annual_salary": sal, "fully_loaded_cost": loaded,
            "is_overhead": None, "revenue_attribution": None,
        } for eid, fn, ln, sup, t, sal, loaded in rows]

        _, _, natural_roots, unreachable = _build_lookups(emp_rows)
        self.assertEqual(unreachable["O1"], "supervisor_not_found")
        self.assertEqual(unreachable["O2"], "extra_root")
        self.assertEqual(unreachable["O3"], "self_referential")
        self.assertEqual(unreachable["C1"], "in_cycle")
        self.assertEqual(unreachable["C2"], "in_cycle")
        self.assertNotIn("K", unreachable)
        self.assertEqual(natural_roots[0], "R")

    def test_revenue_rolls_up_across_three_levels(self):
        """06 §3b: `_agg["revenue_sum"]` reads like a double-counting risk.

        Traced here across three levels with revenue on some nodes and not
        others. It is correct — `revenue_sum` re-derives from the already-summed
        `revenue_managed` for non-leaves, and each parent adds only its own
        `self_revenue` plus its children's sums. Recorded as a test so the next
        person doesn't have to re-derive it.

        metrics.js still leaves revenue to the server (see its docstring); this
        test is what would have to pass before that changes.
        """
        rows = [{
            "employee_id": eid, "first_name": eid, "last_name": "X",
            "raw_supervisor_id": sup, "job_title": "", "management_level": "",
            "department": "", "entity": "", "city": "", "state": "",
            "site_location": "", "employee_type": "", "pay_type": "",
            "annual_salary": None, "fully_loaded_cost": None,
            "is_overhead": None, "revenue_attribution": rev,
        } for eid, sup, rev in [
            ("R",  None, "100"),   # root has its own revenue
            ("M1", "R",  None),    # middle manager has none
            ("M2", "R",  "50"),
            ("L1", "M1", "200"),
            ("L2", "M1", "300"),
            ("L3", "M2", None),    # leaf with none
        ]]
        tree = strip_aggregation(build_tree_from_rows(rows))
        by_id = {}
        stack = [tree]
        while stack:
            n = stack.pop()
            by_id[n["employee_id"]] = n["metrics"]["revenue_managed"]
            stack.extend(n["children"])

        self.assertEqual(by_id["L1"], 200.0)
        self.assertIsNone(by_id["L3"], "a leaf with no revenue reports None, not 0")
        self.assertEqual(by_id["M1"], 500.0, "sum of its two leaves, none of its own")
        self.assertEqual(by_id["M2"], 50.0, "its own only; its leaf has none")
        self.assertEqual(by_id["R"], 650.0, "100 + 500 + 50 — no double counting")

    def test_detect_cycles_is_iterative_on_a_long_chain(self):
        """800 nodes must not blow the stack — this is why it isn't recursive."""
        parent = {f"E{i}": (f"E{i - 1}" if i else None) for i in range(800)}
        self.assertEqual(detect_cycles(parent), [])
        parent["E0"] = "E799"          # close the loop
        cycles = detect_cycles(parent)
        self.assertEqual(len(cycles), 1)
        self.assertEqual(len(cycles[0]), 800)


# ===========================================================================
# Phase 1 — corrections layer and replay
# ===========================================================================

class CorrectionsTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)
        self.user = self.make_user("admin1", self.company, staff=True)

    def _reparent(self, eid, before, after, **kw):
        return self.correction(
            self.company, eid, StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": before}, {"raw_supervisor_id": after}, **kw,
        )

    def test_reparent_correction_applies(self):
        c = self._reparent("E3", "E2", "E5")
        status, _ = C.apply_correction(c, self.snap)
        C.resolve_supervisor_fks(self.snap)
        self.assertEqual(status, StructureCorrection.ReplayStatus.APPLIED)
        emp = Employee.objects.get(snapshot=self.snap, employee_id="E3")
        self.assertEqual(emp.raw_supervisor_id, "E5")
        self.assertEqual(emp.supervisor.employee_id, "E5")

    def test_correction_replays_to_new_snapshot(self):
        self._reparent("E3", "E2", "E5")
        snap_b = self.make_snapshot(self.company, label="Q2")
        log = C.replay_corrections(snap_b, user=self.user)
        self.assertEqual(log.applied_count, 1)
        emp = Employee.objects.get(snapshot=snap_b, employee_id="E3")
        self.assertEqual(emp.raw_supervisor_id, "E5")
        self.assertEqual(emp.supervisor.employee_id, "E5")

    def test_replay_marks_stale_when_person_gone(self):
        c = self._reparent("E3", "E2", "E5")
        rows = [r for r in self.ROWS if r[0] != "E3"]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)
        c.refresh_from_db()
        self.assertEqual(log.stale_count, 1)
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.STALE)
        self.assertFalse(c.is_active)
        snap_b.refresh_from_db()
        self.assertEqual(snap_b.status, CensusSnapshot.Status.ACTIVE)

    def test_replay_marks_conflict_when_target_gone(self):
        c = self._reparent("E3", "E2", "E5")
        rows = [r for r in self.ROWS if r[0] not in ("E5", "E6")]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)
        c.refresh_from_db()
        self.assertEqual(log.conflict_count, 1)
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.CONFLICT)
        self.assertTrue(c.is_active, "a conflict is a to-do, not a retirement")
        self.assertEqual(
            Employee.objects.get(snapshot=snap_b, employee_id="E3").raw_supervisor_id, "E2")

    def test_replay_marks_drifted_when_source_changed(self):
        c = self._reparent("E3", "E2", "E5")
        # The next census says E3 reports to E1 — different from `before` (E2).
        rows = [r if r[0] != "E3" else ("E3", "Cy", "One", "E1", "Technician", "100000", "128000")
                for r in self.ROWS]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)
        c.refresh_from_db()
        self.assertEqual(log.drifted_count, 1)
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.DRIFTED)
        # Drifted still applies — the human decision outranks the export.
        self.assertEqual(
            Employee.objects.get(snapshot=snap_b, employee_id="E3").raw_supervisor_id, "E5")

    def test_replay_rejects_cycle(self):
        """Two reparents that individually validate but jointly cycle."""
        self._reparent("E2", "E1", "E4")   # VP under one of their own reports
        self._reparent("E4", "E2", "E3")   # …which still sits under the VP
        snap_b = self.make_snapshot(self.company, label="Q2")
        log = C.replay_corrections(snap_b)
        self.assertGreaterEqual(log.conflict_count, 1)
        parent = dict(
            Employee.objects.filter(snapshot=snap_b)
            .values_list("employee_id", "raw_supervisor_id"))
        self.assertEqual(detect_cycles(parent), [], "no cycle may survive replay")
        build_tree(snap_b.id)   # must return rather than hang

    def test_revert_restores_original(self):
        c = self._reparent("E3", "E2", "E5")
        C.apply_correction(c, self.snap)
        C.revert_correction(c, self.snap)
        c.refresh_from_db()
        emp = Employee.objects.get(snapshot=self.snap, employee_id="E3")
        self.assertEqual(emp.raw_supervisor_id, "E2")
        self.assertEqual(emp.supervisor.employee_id, "E2")
        self.assertFalse(c.is_active)
        self.assertTrue(StructureCorrection.objects.filter(pk=c.pk).exists())

    def test_exclude_removes_from_tree(self):
        c = self.correction(
            self.company, "E4", StructureCorrection.Kind.EXCLUDE,
            {"employee_status": "", "raw_supervisor_id": "E2"}, {},
        )
        C.apply_correction(c, self.snap)
        tree = strip_aggregation(build_tree(self.snap.id))
        self.assertNotIn("E4", _flatten_ids(tree))
        self.assertTrue(Employee.objects.filter(snapshot=self.snap, employee_id="E4").exists())

    def test_unique_together_updates_not_stacks(self):
        user = self.make_user("editor", self.company, staff=True)
        for target in ("E5", "E1"):
            CS.commit_changeset(
                [{"op": "reparent", "employee_id": "E3",
                  "after": {"raw_supervisor_id": target}, "note": ""}],
                target=CS.TARGET_CORRECTIONS, company=self.company,
                snapshot=self.snap, user=user,
            )
        rows = StructureCorrection.objects.filter(
            company=self.company, employee_id="E3",
            kind=StructureCorrection.Kind.REPARENT)
        self.assertEqual(rows.count(), 1)
        self.assertEqual(rows.first().after["raw_supervisor_id"], "E1")
        # `before` must still be the *source* value, not the intermediate one.
        self.assertEqual(rows.first().before["raw_supervisor_id"], "E2")

    def test_replay_failure_does_not_block_upload(self):
        """A bad correction must never stop a census refresh."""
        from unittest.mock import patch
        self._reparent("E3", "E2", "E5")
        self.make_user("uploader", self.company, staff=True)
        self.client.login(username="uploader", password="pw")

        snap = CensusSnapshot.objects.create(
            company=self.company, label="Q2", original_filename="c.csv",
            status=CensusSnapshot.Status.PROCESSING,
        )
        session = self.client.session
        session["ov_upload"] = {
            "snapshot_id": snap.id, "company_id": self.company.id, "label": "Q2",
            "original_filename": "c.csv", "mapping": {}, "errors": [], "warnings": [],
            "set_as_active": True,
            "mapped_rows": [
                {"employee_id": "E1", "first_name": "Ada", "last_name": "Root",
                 "supervisor_id": "", "job_title": "CEO"},
            ],
        }
        session.save()

        with patch.object(C, "replay_corrections", side_effect=RuntimeError("boom")):
            resp = self.client.post(reverse("org_view:save_snapshot"))
        self.assertEqual(resp.status_code, 302)
        snap.refresh_from_db()
        self.assertEqual(snap.status, CensusSnapshot.Status.ACTIVE)
        self.assertTrue(snap.is_current)

    def _add_person(self, employee_id="51234", first="Mia", last="New", sup="E2"):
        return self.correction(
            self.company, employee_id, StructureCorrection.Kind.ADD_PERSON, {},
            {"first_name": first, "last_name": last, "job_title": "Estimator",
             "raw_supervisor_id": sup},
        )

    def test_add_person_inserts_and_replays(self):
        c = self._add_person()
        status, _ = C.apply_correction(c, self.snap)
        self.assertEqual(status, StructureCorrection.ReplayStatus.APPLIED)

        snap_b = self.make_snapshot(self.company, label="Q2")
        log = C.replay_corrections(snap_b)
        self.assertEqual(log.applied_count, 1)
        emp = Employee.objects.get(snapshot=snap_b, employee_id="51234")
        self.assertEqual(emp.full_name, "Mia New")
        self.assertEqual(emp.supervisor.employee_id, "E2")
        self.assertIn("51234", _flatten_ids(strip_aggregation(build_tree(snap_b.id))))

    def test_add_person_resolves_when_the_export_catches_up(self):
        """The whole point of the resolution rule: no duplicate, and it retires."""
        c = self._add_person()
        C.apply_correction(c, self.snap)

        rows = self.ROWS + [("51234", "Mia", "New", "E2", "Estimator", "90000", None)]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)

        c.refresh_from_db()
        self.assertEqual(log.resolved_count, 1)
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.RESOLVED)
        self.assertFalse(c.is_active, "job done — stop replaying it")
        self.assertEqual(
            Employee.objects.filter(snapshot=snap_b, employee_id="51234").count(), 1)
        # The source row survives untouched — it is real data now.
        self.assertNotIn(
            Employee.ADDED_BY_CORRECTION_KEY,
            Employee.objects.get(snapshot=snap_b, employee_id="51234").raw_data)

    def test_add_person_resolves_on_a_name_match_under_a_different_badge(self):
        """Guards the case where the user had to invent an id.

        06 §6 flags exactly this: a generated id can't match the export, so we
        fall back to the name before creating a second copy of a real person.
        """
        c = self._add_person(employee_id="ADDED-1")
        C.apply_correction(c, self.snap)

        rows = self.ROWS + [("77777", "Mia", "New", "E2", "Estimator", "90000", None)]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)

        c.refresh_from_db()
        self.assertEqual(log.resolved_count, 1)
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.RESOLVED)
        self.assertFalse(
            Employee.objects.filter(snapshot=snap_b, employee_id="ADDED-1").exists())
        self.assertEqual(
            Employee.objects.filter(snapshot=snap_b, last_name="New").count(), 1)

    def test_add_person_conflicts_when_the_manager_is_gone(self):
        c = self._add_person(sup="E5")
        C.apply_correction(c, self.snap)
        rows = [r for r in self.ROWS if r[0] not in ("E5", "E6")]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        log = C.replay_corrections(snap_b)
        c.refresh_from_db()
        self.assertEqual(log.conflict_count, 1)
        self.assertTrue(c.is_active)
        self.assertFalse(
            Employee.objects.filter(snapshot=snap_b, employee_id="51234").exists())

    def test_revert_add_person_deletes_our_row_and_reparents_its_reports(self):
        c = self._add_person()
        C.apply_correction(c, self.snap)
        C.resolve_supervisor_fks(self.snap)
        # Someone was moved under the person we added.
        Employee.objects.filter(snapshot=self.snap, employee_id="E4").update(
            raw_supervisor_id="51234")

        C.revert_correction(c, self.snap)
        c.refresh_from_db()
        self.assertFalse(c.is_active)
        self.assertFalse(
            Employee.objects.filter(snapshot=self.snap, employee_id="51234").exists())
        # Their report must not be orphaned by the revert.
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E4").raw_supervisor_id, "E2")

    def test_revert_add_person_never_deletes_a_real_source_row(self):
        c = self._add_person()
        rows = self.ROWS + [("51234", "Mia", "New", "E2", "Estimator", "90000", None)]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        C.revert_correction(c, snap_b)
        self.assertTrue(
            Employee.objects.filter(snapshot=snap_b, employee_id="51234").exists(),
            "the export owns this row now — reverting must not destroy real data")

    def test_management_command_dry_run(self):
        self._reparent("E3", "E2", "E5")
        call_command("replay_corrections", company="Flooring Partners", dry_run=True)
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E3").raw_supervisor_id, "E2")
        self.assertFalse(CorrectionReplayLog.objects.exists())


# ===========================================================================
# Phase 2 — the changeset API
# ===========================================================================

class ChangesetTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)
        self.admin = self.make_user("admin2", self.company, role="admin")
        self.client.login(username="admin2", password="pw")

    def url(self, name, **kw):
        return reverse(f"org_view:{name}", kwargs={"slug": self.company.slug, **kw})

    def post(self, name, payload, **kw):
        return self.client.post(
            self.url(name, **kw), data=json.dumps(payload), content_type="application/json")

    def commit(self, ops, **extra):
        return self.post("api_commit_changeset",
                         {"target": "corrections", "ops": ops, **extra})

    def test_batch_cycle_rejected(self):
        """Two individually-valid reparents that jointly cycle. The headline test."""
        before = dict(Employee.objects.filter(snapshot=self.snap)
                      .values_list("employee_id", "raw_supervisor_id"))
        resp = self.commit([
            {"op": "reparent", "employee_id": "E2", "after": {"raw_supervisor_id": "E3"}},
            {"op": "reparent", "employee_id": "E3", "after": {"raw_supervisor_id": "E2"}},
        ])
        self.assertEqual(resp.status_code, 422)
        self.assertTrue(any("loop" in e["error"] for e in resp.json()["errors"]))
        after = dict(Employee.objects.filter(snapshot=self.snap)
                     .values_list("employee_id", "raw_supervisor_id"))
        self.assertEqual(before, after, "a rejected batch must write nothing")
        self.assertFalse(StructureCorrection.objects.exists())

    def test_atomic_rollback(self):
        resp = self.commit([
            {"op": "attribute", "employee_id": "E3", "after": {"job_title": "Lead Tech"}},
            {"op": "attribute", "employee_id": "E4", "after": {"department": "Service"}},
            {"op": "reparent",  "employee_id": "E6", "after": {"raw_supervisor_id": "E2"}},
            {"op": "reparent",  "employee_id": "E4", "after": {"raw_supervisor_id": "NOPE"}},
            {"op": "attribute", "employee_id": "E1", "after": {"city": "Tacoma"}},
        ])
        self.assertEqual(resp.status_code, 422)
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E3").job_title, "Technician")
        self.assertEqual(StructureCorrection.objects.count(), 0)

    def test_unknown_field_is_error_not_silent_drop(self):
        resp = self.commit([
            {"op": "attribute", "employee_id": "E3", "after": {"bogus_field": "x"}}])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("bogus_field", resp.json()["errors"][0]["error"])

    def test_correct_mode_eliminate_pulls_reports_up(self):
        """Eliminate behaves the same as it does in a scenario."""
        resp = self.commit([{"op": "eliminate", "employee_id": "E2", "after": {},
                             "note": "role closed in April"}])
        self.assertEqual(resp.status_code, 200, resp.content)

        gone = Employee.objects.get(snapshot=self.snap, employee_id="E2")
        self.assertEqual(gone.employee_status, Employee.EXCLUDED_STATUS)
        self.assertNotIn("E2", _flatten_ids(strip_aggregation(build_tree(self.snap.id))))
        # …and their team is not stranded.
        for eid in ("E3", "E4"):
            child = Employee.objects.get(snapshot=self.snap, employee_id=eid)
            self.assertEqual(child.raw_supervisor_id, "E1")
            self.assertEqual(child.supervisor.employee_id, "E1")
        self.assertEqual(self.client.get(self.url("api_unattached")).json()["counts"]["orphans"], 1,
                         "only the pre-existing X1 orphan — no new ones")

    def test_removal_allocates_reports_individually(self):
        """A disbanded team usually splits across several managers."""
        resp = self.commit([{
            "op": "eliminate", "employee_id": "E2",
            "after": {"reassign": {"E3": "E5"}},   # E4 left blank -> falls back
            "note": "team split between ops and finance",
        }])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E3").raw_supervisor_id, "E5")
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E4").raw_supervisor_id, "E1",
            "an unallocated report falls back to the removed manager's own manager")

    def test_removal_allocation_replays_to_the_next_census(self):
        self.commit([{"op": "eliminate", "employee_id": "E2",
                      "after": {"reassign": {"E3": "E5"}}}])
        snap_b = self.make_snapshot(self.company, label="Q2")
        C.replay_corrections(snap_b)
        self.assertEqual(
            Employee.objects.get(snapshot=snap_b, employee_id="E3").raw_supervisor_id, "E5")
        self.assertEqual(
            Employee.objects.get(snapshot=snap_b, employee_id="E4").raw_supervisor_id, "E1")
        self.assertNotIn("E2", _flatten_ids(strip_aggregation(build_tree(snap_b.id))))

    def test_removal_rejects_an_allocation_into_the_removed_subtree(self):
        for reassign, fragment in (
            ({"E3": "E2"}, "position being removed"),
            ({"E3": "E3"}, "their own manager"),
            ({"E3": "NOPE"}, "not in this org"),
            ({"E6": "E5"}, "doesn't report to"),
        ):
            resp = self.commit([{"op": "eliminate", "employee_id": "E2",
                                 "after": {"reassign": reassign}}])
            self.assertEqual(resp.status_code, 422, reassign)
            self.assertIn(fragment, resp.json()["errors"][0]["error"])
        self.assertFalse(StructureCorrection.objects.exists())

    def test_reverting_a_removal_puts_only_its_own_reports_back(self):
        self.commit([{"op": "eliminate", "employee_id": "E2",
                      "after": {"reassign": {"E3": "E5"}}}])
        correction = StructureCorrection.objects.get(kind=StructureCorrection.Kind.ELIMINATE)
        self.client.post(
            reverse("org_view:api_revert_correction",
                    kwargs={"slug": self.company.slug, "pk": correction.pk}),
            data="{}", content_type="application/json")

        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E2").raw_supervisor_id, "E1")
        for eid in ("E3", "E4"):
            self.assertEqual(
                Employee.objects.get(snapshot=self.snap, employee_id=eid).raw_supervisor_id,
                "E2", "both reports go back under the restored manager")
        # E5's own report was never E2's and must not have been swept up.
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E6").raw_supervisor_id, "E5")

    def test_exclude_no_longer_strands_the_excluded_manager_team(self):
        """Regression: excluding a manager used to drop their whole team into
        the orphan tray, because the reports kept pointing at a filtered id."""
        before = self.client.get(self.url("api_unattached")).json()["counts"]["orphans"]
        resp = self.commit([{"op": "exclude", "employee_id": "E2", "after": {}}])
        self.assertEqual(resp.status_code, 200, resp.content)
        after = resp.json()["summary"]["counts"]["orphans"]
        self.assertEqual(after, before, "no new orphans")
        for eid in ("E3", "E4"):
            self.assertIn(eid, _flatten_ids(strip_aggregation(build_tree(self.snap.id))))

    def test_scenario_eliminate_honours_allocation(self):
        scenario = S.create_scenario(company=self.company, base_snapshot=self.snap, name="R")
        resp = self.post("api_commit_changeset", {
            "target": "scenario", "scenario_id": scenario.id,
            "ops": [{"op": "eliminate", "employee_id": "E2",
                     "after": {"reassign": {"E3": "E5"}}}]})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(scenario.positions.get(employee_id="E3").raw_supervisor_id, "E5")
        self.assertEqual(scenario.positions.get(employee_id="E4").raw_supervisor_id, "E1")
        self.assertEqual(scenario.positions.get(employee_id="E3").prior_supervisor_id, "E2")

    def test_scenario_mode_still_rejects_exclude(self):
        """A duplicate row is a statement about the census, not about a reorg."""
        scenario = S.create_scenario(company=self.company, base_snapshot=self.snap, name="R")
        resp = self.post("api_commit_changeset", {
            "target": "scenario", "scenario_id": scenario.id,
            "ops": [{"op": "exclude", "employee_id": "E3", "after": {}}]})
        self.assertEqual(resp.status_code, 422)
        self.assertIn("eliminate the position instead", resp.json()["errors"][0]["error"])

    def test_correct_mode_allows_add(self):
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"employee_id": "51234", "first_name": "Mia", "last_name": "New",
                      "job_title": "Estimator", "raw_supervisor_id": "E2"},
            "note": "missing from the March export",
        }])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id_map"], {"TMP-1": "51234"})
        emp = Employee.objects.get(snapshot=self.snap, employee_id="51234")
        self.assertEqual(emp.full_name, "Mia New")
        self.assertEqual(emp.supervisor.employee_id, "E2")
        self.assertTrue(emp.raw_data[Employee.ADDED_BY_CORRECTION_KEY])

    def test_correct_mode_add_rejects_a_duplicate_badge(self):
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"employee_id": "E3", "first_name": "Imp", "last_name": "Ostor",
                      "raw_supervisor_id": "E2"},
        }])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("already in this census", resp.json()["errors"][0]["error"])

    def test_correct_mode_add_requires_a_name_or_title(self):
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"employee_id": "51234", "raw_supervisor_id": "E2"},
        }])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("name or a job title", resp.json()["errors"][0]["error"])

    def test_correct_mode_add_generates_an_id_when_none_given(self):
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"first_name": "No", "last_name": "Badge", "raw_supervisor_id": "E2"},
        }])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id_map"], {"TMP-1": "ADDED-1"})

    def test_correct_mode_add_rejects_pay_fields(self):
        """Pay comes from the payroll export, not from someone typing it in."""
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"employee_id": "51234", "first_name": "Mia", "last_name": "New",
                      "annual_salary": "90000", "raw_supervisor_id": "E2"},
        }])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("annual_salary", resp.json()["errors"][0]["error"])

    def test_add_then_reparent_someone_under_them_in_one_batch(self):
        resp = self.commit([
            {"op": "add", "employee_id": "TMP-1",
             "after": {"employee_id": "51234", "first_name": "Mia", "last_name": "New",
                       "job_title": "Ops Director", "raw_supervisor_id": "E1"}},
            {"op": "reparent", "employee_id": "E2", "after": {"raw_supervisor_id": "TMP-1"}},
        ])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E2").raw_supervisor_id,
            "51234", "the reparent must resolve the temp id to the assigned badge")

    def test_add_that_the_census_already_covers_remaps_its_reports(self):
        """An add can stand down — and its reports still have to land somewhere.

        `_apply_add_person` retires the correction when the census already has this
        person under a different badge number: ours would duplicate them. Nothing
        is created at the id we asked for, so anyone the user put under the new
        role must be pointed at the row that does exist. Writing the unused id
        would file them under a supervisor `_fetch_employees` never returns, i.e.
        silently off the chart, on a save that reported success.
        """
        resp = self.commit([
            {"op": "add", "employee_id": "TMP-1",
             "after": {"employee_id": "51234", "first_name": "Cy", "last_name": "One",
                       "job_title": "Technician", "raw_supervisor_id": "E1"}},
            {"op": "reparent", "employee_id": "E4", "after": {"raw_supervisor_id": "TMP-1"}},
        ])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id_map"], {"TMP-1": "E3"},
                         "the temp id must resolve to the row that actually exists")
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E4").raw_supervisor_id,
            "E3")
        self.assertFalse(
            Employee.objects.filter(snapshot=self.snap, employee_id="51234").exists())
        self.assertIn("E4", _flatten_ids(build_tree(self.snap.id)),
                      "the reassigned report must still be drawn")

    def test_add_with_nowhere_to_land_fails_the_batch_instead_of_orphaning(self):
        """When an add materializes nothing at all, dependents can't be written.

        Here the only name match is a row that was excluded from the chart, so
        `_apply_add_person` stands down and there is no live row to point at. The
        batch is refused whole rather than stranding E4 under an id that names
        nobody.
        """
        self.assertEqual(
            self.commit([{"op": "exclude", "employee_id": "E3", "after": {}}]).status_code,
            200)
        before = dict(Employee.objects.filter(snapshot=self.snap)
                      .values_list("employee_id", "raw_supervisor_id"))

        resp = self.commit([
            {"op": "add", "employee_id": "TMP-1",
             "after": {"employee_id": "51234", "first_name": "Cy", "last_name": "One",
                       "job_title": "Technician", "raw_supervisor_id": "E1"}},
            {"op": "reparent", "employee_id": "E4", "after": {"raw_supervisor_id": "TMP-1"}},
        ])
        self.assertEqual(resp.status_code, 422, resp.content)
        error = resp.json()["errors"][0]["error"]
        self.assertIn("would duplicate them", error, "say why it stood down")
        self.assertIn("nowhere to go", error, "and what that means for the batch")
        self.assertEqual(
            before,
            dict(Employee.objects.filter(snapshot=self.snap)
                 .values_list("employee_id", "raw_supervisor_id")),
            "a refused batch must write nothing")

    def test_an_add_nothing_depends_on_still_commits(self):
        """The guard fires only when something would be left dangling."""
        self.assertEqual(
            self.commit([{"op": "exclude", "employee_id": "E3", "after": {}}]).status_code,
            200)
        resp = self.commit([{
            "op": "add", "employee_id": "TMP-1",
            "after": {"employee_id": "51234", "first_name": "Cy", "last_name": "One",
                      "job_title": "Technician", "raw_supervisor_id": "E1"},
        }])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id_map"], {},
                         "nothing landed, so there is no id to map")
        correction = StructureCorrection.objects.get(
            employee_id="51234", kind=StructureCorrection.Kind.ADD_PERSON)
        self.assertEqual(correction.replay_status,
                         StructureCorrection.ReplayStatus.RESOLVED)

    def test_correct_mode_rejects_pay_fields(self):
        resp = self.commit([
            {"op": "attribute", "employee_id": "E3", "after": {"annual_salary": "1"}}])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("annual_salary", resp.json()["errors"][0]["error"])

    def test_restricted_role_cannot_set_pay(self):
        self.make_user("restricted1", self.company, role="restricted")
        scenario = S.create_scenario(company=self.company, base_snapshot=self.snap, name="R")
        self.client.logout()
        self.client.login(username="restricted1", password="pw")
        resp = self.post("api_commit_changeset", {
            "target": "scenario", "scenario_id": scenario.id,
            "ops": [{"op": "attribute", "employee_id": "E3",
                     "after": {"annual_salary": "1"}}],
        })
        # A restricted user isn't an edit role at all, so it never reaches the
        # whitelist — but the validator must refuse it independently too, since
        # the old form only avoided this by accident (no pay inputs rendered).
        self.assertEqual(resp.status_code, 403)
        errors = CS.validate_changeset(
            [{"op": "attribute", "employee_id": "E3", "after": {"annual_salary": "1"}}],
            target=CS.TARGET_SCENARIO, scenario=scenario,
            whitelist=S.EDITABLE_FIELDS, can_see_pay=False)
        self.assertTrue(any("pay" in e["error"] for e in errors))

    def test_branch_admin_cannot_reparent_outside_branch(self):
        self.make_user("branchadmin", self.company, role="admin", branch="E2")
        self.client.logout()
        self.client.login(username="branchadmin", password="pw")
        resp = self.commit([
            {"op": "reparent", "employee_id": "E3", "after": {"raw_supervisor_id": "E5"}}])
        self.assertEqual(resp.status_code, 422)
        self.assertIn("outside", resp.json()["errors"][0]["error"])

    def test_stale_snapshot_returns_409(self):
        newer = self.make_snapshot(self.company, label="Q2")
        resp = self.commit(
            [{"op": "attribute", "employee_id": "E3", "after": {"job_title": "X"}}],
            expected_snapshot_id=self.snap.id)
        self.assertEqual(resp.status_code, 409)
        self.assertEqual(resp.json()["current_snapshot_id"], newer.id)
        self.assertFalse(StructureCorrection.objects.exists())

    def test_temp_id_mapping(self):
        scenario = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Plan")
        resp = self.post("api_commit_changeset", {
            "target": "scenario", "scenario_id": scenario.id,
            "ops": [
                {"op": "add", "employee_id": "TMP-1",
                 "after": {"raw_supervisor_id": "E1", "job_title": "VP Finance",
                           "is_vacant": True}},
                {"op": "reparent", "employee_id": "E5",
                 "after": {"raw_supervisor_id": "TMP-1"}},
            ],
        })
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id_map"], {"TMP-1": "NEW-1"})
        self.assertEqual(scenario.positions.get(employee_id="E5").raw_supervisor_id, "NEW-1")

    def test_commit_returns_fresh_tree_and_summary(self):
        resp = self.commit([
            {"op": "reparent", "employee_id": "E6", "after": {"raw_supervisor_id": "E2"}}])
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        e2 = next(c for c in body["tree"]["children"] if c["employee_id"] == "E2")
        self.assertIn("E6", [c["employee_id"] for c in e2["children"]])
        fresh = strip_aggregation(build_tree(self.snap.id))
        self.assertEqual(body["summary"]["counts"]["rendered"], len(_flatten_ids(fresh)))

    def test_unattached_lists_orphan_with_missing_supervisor(self):
        resp = self.client.get(self.url("api_unattached"))
        self.assertEqual(resp.status_code, 200)
        x1 = next(o for o in resp.json()["orphans"] if o["employee_id"] == "X1")
        self.assertEqual(x1["reason"], "supervisor_not_found")
        self.assertEqual(x1["raw_supervisor_id"], "9999")
        self.assertNotIn("X1", _flatten_ids(strip_aggregation(build_tree(self.snap.id))))

    def test_unattached_counts_reconcile(self):
        c = self.correction(
            self.company, "E4", StructureCorrection.Kind.EXCLUDE,
            {"employee_status": "", "raw_supervisor_id": "E2"}, {})
        C.apply_correction(c, self.snap)
        counts = self.client.get(self.url("api_unattached")).json()["counts"]
        self.assertEqual(
            counts["rendered"] + counts["orphans"] + counts["excluded"],
            counts["total_employees"])
        self.assertEqual(counts["excluded"], 1)

    def test_commit_redacts_pay_for_restricted(self):
        self.make_user("restricted2", self.company, role="restricted")
        self.client.logout()
        self.client.login(username="restricted2", password="pw")
        tree = self.client.get(self.url("api_org_tree")).json()["tree"]
        self.assertIsNone(tree["metrics"]["total_labor_cost"])
        self.assertIsNone(tree["self"]["cost"])

    def test_validate_and_commit_reject_identically(self):
        ops = [
            {"op": "reparent", "employee_id": "E2", "after": {"raw_supervisor_id": "E3"}},
            {"op": "reparent", "employee_id": "E3", "after": {"raw_supervisor_id": "E2"}},
        ]
        v = self.post("api_validate_changeset", {"target": "corrections", "ops": ops})
        c = self.commit(ops)
        self.assertEqual(v.status_code, 200)
        self.assertFalse(v.json()["valid"])
        self.assertEqual(c.status_code, 422)
        self.assertEqual([e["error"] for e in v.json()["errors"]],
                         [e["error"] for e in c.json()["errors"]])

    def test_non_admin_gets_403_on_every_editing_endpoint(self):
        self.make_user("viewer1", self.company, role="viewer")
        self.client.logout()
        self.client.login(username="viewer1", password="pw")
        for name in ("api_validate_changeset", "api_commit_changeset"):
            self.assertEqual(
                self.post(name, {"target": "corrections", "ops": []}).status_code, 403, name)
        c = self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"})
        resp = self.client.post(
            reverse("org_view:api_revert_correction",
                    kwargs={"slug": self.company.slug, "pk": c.pk}),
            data="{}", content_type="application/json")
        self.assertEqual(resp.status_code, 403)
        # …but the read endpoints stay open to any permission.
        self.assertEqual(self.client.get(self.url("api_unattached")).status_code, 200)
        self.assertEqual(self.client.get(self.url("api_corrections")).status_code, 200)

    def test_unattached_orphan_carries_its_whole_cluster(self):
        """An orphan can be a manager whose team is equally invisible.

        The tray lists the cluster root once and ships its members, so attaching
        the manager brings the team instead of silently stranding it.
        """
        for eid, sup in (("X1a", "X1"), ("X1b", "X1"), ("X1c", "X1a")):
            Employee.objects.create(
                snapshot=self.snap, employee_id=eid, first_name=eid, last_name="Team",
                raw_supervisor_id=sup, job_title="Installer")
        body = self.client.get(self.url("api_unattached")).json()

        roots = [o["employee_id"] for o in body["orphans"]]
        self.assertEqual(roots, ["X1"], "only the cluster root is listed")
        x1 = body["orphans"][0]
        self.assertEqual(x1["subtree_count"], 3)
        self.assertEqual({p["employee_id"] for p in x1["cluster"]}, {"X1a", "X1b", "X1c"})
        # counts still reconcile even though the list has one row for four people
        c = body["counts"]
        self.assertEqual(c["orphans"], 4)
        self.assertEqual(c["rendered"] + c["orphans"] + c["excluded"], c["total_employees"])

    def test_employee_raw_endpoint_redacts_pay_for_restricted(self):
        emp = Employee.objects.get(snapshot=self.snap, employee_id="E3")
        emp.raw_data = {"employee_id": "E3", "annual_salary": "100000", "dept": "Ops"}
        emp.save(update_fields=["raw_data"])
        url = reverse("org_view:api_employee_raw",
                      kwargs={"slug": self.company.slug, "employee_id": "E3"})

        raw = self.client.get(url).json()["raw_data"]
        self.assertIn("annual_salary", raw)

        self.make_user("restricted5", self.company, role="restricted")
        self.client.logout()
        self.client.login(username="restricted5", password="pw")
        raw = self.client.get(url).json()["raw_data"]
        self.assertNotIn("annual_salary", raw)
        self.assertIn("dept", raw)

    def test_oversized_batch_returns_413(self):
        ops = [{"op": "attribute", "employee_id": "E3", "after": {"city": str(i)}}
               for i in range(CS.MAX_OPS + 1)]
        self.assertEqual(self.commit(ops).status_code, 413)


# ===========================================================================
# Exports
# ===========================================================================

class ExportTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)
        self.make_user("admin7", self.company, role="admin")
        self.client.login(username="admin7", password="pw")

    def url(self, name, **kw):
        return reverse(f"org_view:{name}", kwargs={"slug": self.company.slug, **kw})

    def csv_rows(self, response):
        text = response.content.decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))

    def commit(self, ops):
        return self.client.post(
            self.url("api_commit_changeset"),
            data=json.dumps({"target": "corrections", "ops": ops}),
            content_type="application/json")

    # ── The corrected census ────────────────────────────────────────

    def test_census_csv_round_trips_through_the_importer(self):
        """The file we hand back must be re-importable without remapping."""
        from .parsers import auto_map_columns
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        headers = rows[0]
        mapping = auto_map_columns([h for h in headers if not h.startswith("_")])
        for required in ("employee_id", "first_name", "last_name", "supervisor_id"):
            self.assertIsNotNone(mapping.get(required), f"{required} would not auto-map")
        self.assertEqual(len(rows) - 1, 7)

    def test_census_export_reflects_corrections_not_the_raw_import(self):
        self.commit([{"op": "reparent", "employee_id": "E3",
                      "after": {"raw_supervisor_id": "E5"}, "note": "moved in April"}])
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        headers = rows[0]
        row = next(r for r in rows[1:] if r[headers.index("employee_id")] == "E3")
        self.assertEqual(row[headers.index("supervisor_id")], "E5")
        self.assertEqual(row[headers.index("_corrections")], "Change manager")

    def test_census_export_omits_pay_columns_for_restricted(self):
        """Omitted, not blanked — a blank column still says the field exists."""
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        self.assertIn("annual_salary", rows[0])

        self.make_user("restricted6", self.company, role="restricted")
        self.client.logout()
        self.client.login(username="restricted6", password="pw")
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        for column in ("annual_salary", "fully_loaded_cost", "revenue_attribution"):
            self.assertNotIn(column, rows[0])
        self.assertNotIn(b"300000", self.client.get(
            self.url("export_census") + "?format=csv").content)

    def test_census_export_is_branch_scoped(self):
        self.make_user("branch5", self.company, role="admin", branch="E2")
        self.client.logout()
        self.client.login(username="branch5", password="pw")
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        ids = {r[rows[0].index("employee_id")] for r in rows[1:]}
        self.assertEqual(ids, {"E2", "E3", "E4"})

    def test_census_export_omits_excluded_rows_unless_asked(self):
        self.commit([{"op": "exclude", "employee_id": "E4", "after": {}}])
        rows = self.csv_rows(self.client.get(self.url("export_census") + "?format=csv"))
        ids = {r[rows[0].index("employee_id")] for r in rows[1:]}
        self.assertNotIn("E4", ids)
        rows = self.csv_rows(self.client.get(
            self.url("export_census") + "?format=csv&include_excluded=1"))
        ids = {r[rows[0].index("employee_id")] for r in rows[1:]}
        self.assertIn("E4", ids)

    def test_census_xlsx_carries_data_actions_and_ledger(self):
        import openpyxl
        self.commit([{"op": "reparent", "employee_id": "E3",
                      "after": {"raw_supervisor_id": "E5"}}])
        resp = self.client.get(self.url("export_census"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertEqual(wb.sheetnames, ["Census", "HR actions", "Corrections"])

    # ── The HR action list ──────────────────────────────────────────

    def test_hr_actions_lists_one_row_per_change(self):
        self.commit([
            {"op": "reparent", "employee_id": "E3",
             "after": {"raw_supervisor_id": "E5"}, "note": "moved to finance"},
            {"op": "attribute", "employee_id": "E4",
             "after": {"job_title": "Senior Technician", "department": "Service"}},
        ])
        rows = self.csv_rows(self.client.get(self.url("export_hr_actions")))
        headers = rows[0]
        self.assertIn("current value in HR system", headers)
        self.assertIn("change to", headers)

        body = rows[1:]
        # One row for the move, and one per corrected field — a tickable list.
        self.assertEqual(len(body), 3)
        move = next(r for r in body if r[headers.index("employee_id")] == "E3")
        self.assertEqual(move[headers.index("action")], "Change manager")
        self.assertIn("Bo Vice (E2)", move[headers.index("current value in HR system")])
        self.assertIn("El Three (E5)", move[headers.index("change to")])
        self.assertEqual(move[headers.index("why")], "moved to finance")

        fields = {r[headers.index("field")] for r in body
                  if r[headers.index("employee_id")] == "E4"}
        self.assertEqual(fields, {"Job Title", "Department"})

    def test_hr_actions_includes_reports_moved_by_an_elimination(self):
        """Eliminating a manager is a manager change for every one of their reports."""
        self.commit([{"op": "eliminate", "employee_id": "E2",
                      "after": {"reassign": {"E3": "E5"}}, "note": "role closed"}])
        rows = self.csv_rows(self.client.get(self.url("export_hr_actions")))
        headers = rows[0]
        body = rows[1:]

        removal = next(r for r in body if r[headers.index("employee_id")] == "E2")
        self.assertEqual(removal[headers.index("action")], "Remove position")

        moves = {r[headers.index("employee_id")]: r for r in body
                 if r[headers.index("action")] == "Change manager"}
        self.assertEqual(set(moves), {"E3", "E4"})
        self.assertIn("El Three (E5)", moves["E3"][headers.index("change to")])
        self.assertIn("Ada Root (E1)", moves["E4"][headers.index("change to")])
        self.assertIn("eliminated", moves["E3"][headers.index("why")])

    def test_hr_actions_flags_what_the_source_has_already_done(self):
        self.commit([{"op": "reparent", "employee_id": "E3",
                      "after": {"raw_supervisor_id": "E5"}}])
        # Next census already reports E3 under someone else — the export moved on.
        rows = [r if r[0] != "E3" else ("E3", "Cy", "One", "E1", "Technician", "100000", "128000")
                for r in self.ROWS]
        snap_b = self.make_snapshot(self.company, rows=rows, label="Q2")
        C.replay_corrections(snap_b)

        out = self.csv_rows(self.client.get(self.url("export_hr_actions")))
        headers, body = out[0], out[1:]
        row = next(r for r in body if r[headers.index("employee_id")] == "E3")
        self.assertTrue(row[headers.index("status")].startswith("Check first"))

    def test_hr_actions_sorts_live_work_first(self):
        self.commit([{"op": "reparent", "employee_id": "E3",
                      "after": {"raw_supervisor_id": "E5"}}])
        done = self.correction(
            self.company, "E4", StructureCorrection.Kind.ATTRIBUTE,
            {"job_title": "Technician"}, {"job_title": "Lead"},
            replay_status=StructureCorrection.ReplayStatus.RESOLVED)
        self.assertTrue(done.pk)
        out = self.csv_rows(self.client.get(self.url("export_hr_actions")))
        headers, body = out[0], out[1:]
        self.assertEqual(body[0][headers.index("status")].split(" —")[0], "Action needed")

    def test_hr_actions_hides_reverted_corrections_unless_asked(self):
        self.commit([{"op": "reparent", "employee_id": "E3",
                      "after": {"raw_supervisor_id": "E5"}}])
        correction = StructureCorrection.objects.get()
        C.revert_correction(correction, self.snap)
        self.assertEqual(len(self.csv_rows(self.client.get(self.url("export_hr_actions")))), 1)
        rows = self.csv_rows(self.client.get(
            self.url("export_hr_actions") + "?include_inactive=1"))
        self.assertEqual(len(rows), 2)

    # ── Scenarios ───────────────────────────────────────────────────

    def test_scenario_export_has_three_sheets_and_hides_pay(self):
        import openpyxl
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="FY27")
        S.eliminate_position(sc.positions.get(employee_id="E2"))
        url = reverse("org_view:export_scenario",
                      kwargs={"slug": self.company.slug, "scenario_id": sc.id})

        wb = openpyxl.load_workbook(io.BytesIO(self.client.get(url).content))
        self.assertEqual(wb.sheetnames, ["Impact", "Change log", "Positions"])
        self.assertIn("cost_impact", [c.value for c in wb["Change log"][1]])

        self.make_user("restricted7", self.company, role="restricted")
        self.client.logout()
        self.client.login(username="restricted7", password="pw")
        wb = openpyxl.load_workbook(io.BytesIO(self.client.get(url).content))
        self.assertNotIn("cost_impact", [c.value for c in wb["Change log"][1]])
        self.assertNotIn("annual_salary", [c.value for c in wb["Positions"][1]])

    def test_exports_require_company_access(self):
        other = Company.objects.create(name="Other", slug="other")
        CensusSnapshot.objects.create(
            company=other, original_filename="x.csv",
            status=CensusSnapshot.Status.ACTIVE, is_current=True)
        for name in ("export_census", "export_hr_actions", "export_corrections"):
            resp = self.client.get(
                reverse(f"org_view:{name}", kwargs={"slug": "other"}))
            self.assertEqual(resp.status_code, 302, name)


# ===========================================================================
# Decorative grouping boxes
# ===========================================================================

class ChartGroupTests(OrgFixtureMixin, TestCase):
    """Grouping is presentational: it must change nothing about the data."""

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)
        self.make_user("admin6", self.company, role="admin")
        self.client.login(username="admin6", password="pw")

    def url(self, name, **kw):
        return reverse(f"org_view:{name}", kwargs={"slug": self.company.slug, **kw})

    def save(self, **body):
        payload = {"parent_employee_id": "E2", "name": "Field Service",
                   "member_ids": ["E3", "E4"], "accent": "sage"}
        payload.update(body)
        return self.client.post(self.url("api_save_chart_group"),
                                data=json.dumps(payload), content_type="application/json")

    def test_group_round_trips_and_changes_no_reporting_line(self):
        before = dict(Employee.objects.filter(snapshot=self.snap)
                      .values_list("employee_id", "raw_supervisor_id"))
        resp = self.save()
        self.assertEqual(resp.status_code, 200, resp.content)

        groups = self.client.get(self.url("api_chart_groups")).json()["groups"]
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]["name"], "Field Service")
        self.assertEqual(groups[0]["member_ids"], ["E3", "E4"])
        self.assertEqual(groups[0]["accent"], "sage")

        after = dict(Employee.objects.filter(snapshot=self.snap)
                     .values_list("employee_id", "raw_supervisor_id"))
        self.assertEqual(before, after, "a group must not touch the structure")
        # …and it is not a correction.
        self.assertFalse(StructureCorrection.objects.exists())

    def test_group_survives_a_census_reupload(self):
        """Keyed by employee_id, like corrections, so cleanup happens once."""
        self.save()
        snap_b = self.make_snapshot(self.company, label="Q2")
        groups = self.client.get(self.url("api_chart_groups")).json()["groups"]
        self.assertEqual(groups[0]["member_ids"], ["E3", "E4"])
        self.assertTrue(Employee.objects.filter(snapshot=snap_b, employee_id="E3").exists())

    def test_group_rejects_an_empty_name_or_membership(self):
        self.assertEqual(self.save(name="  ").status_code, 400)
        self.assertEqual(self.save(member_ids=[]).status_code, 400)
        self.assertFalse(ChartGroup.objects.exists())

    def test_group_name_must_be_unique_per_company(self):
        self.assertEqual(self.save().status_code, 200)
        self.assertEqual(
            self.save(parent_employee_id="E1", member_ids=["E5"]).status_code, 409)

    def test_membership_is_free_form_and_needs_no_anchor(self):
        """Members need not share a manager, and the anchor is optional.

        E3/E4 report to E2 and E6 reports to E5 — a box can still hold all three,
        and the client works out where to hang it.
        """
        resp = self.save(parent_employee_id="", member_ids=["E3", "E4", "E6"])
        self.assertEqual(resp.status_code, 200, resp.content)
        g = ChartGroup.objects.get()
        self.assertEqual(g.parent_employee_id, "")
        self.assertEqual(g.member_ids, ["E3", "E4", "E6"])
        # Still nothing structural has moved.
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E6").raw_supervisor_id, "E5")

    def test_a_person_can_only_be_in_one_box(self):
        self.assertEqual(self.save().status_code, 200)
        resp = self.save(name="Another team", member_ids=["E4", "E6"])
        self.assertEqual(resp.status_code, 409)
        self.assertIn("already belongs to", resp.json()["error"])

    def test_group_rejects_members_missing_from_the_census(self):
        resp = self.save(member_ids=["E3", "GHOST"])
        self.assertEqual(resp.status_code, 400)
        self.assertIn("GHOST", resp.json()["error"])

    def test_group_cannot_be_placed_under_one_of_its_own_members(self):
        resp = self.save(parent_employee_id="E3", member_ids=["E3", "E4"])
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(ChartGroup.objects.exists())

    def test_explicit_anchor_is_kept(self):
        resp = self.save(parent_employee_id="E1", member_ids=["E3", "E4"])
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(ChartGroup.objects.get().parent_employee_id, "E1")

    def test_updating_a_group_replaces_its_membership(self):
        gid = self.save().json()["id"]
        resp = self.save(id=gid, member_ids=["E3"], name="Service", accent="plum")
        self.assertEqual(resp.status_code, 200)
        g = ChartGroup.objects.get(pk=gid)
        self.assertEqual(g.member_ids, ["E3"])
        self.assertEqual(g.name, "Service")
        self.assertEqual(g.accent, "plum")
        self.assertEqual(ChartGroup.objects.count(), 1)

    def test_ungrouping_removes_only_the_box(self):
        gid = self.save().json()["id"]
        resp = self.client.post(
            self.url("api_delete_chart_group", pk=gid), data="{}",
            content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(ChartGroup.objects.exists())
        for eid in ("E3", "E4"):
            self.assertEqual(
                Employee.objects.get(snapshot=self.snap, employee_id=eid).raw_supervisor_id,
                "E2", "members keep reporting exactly where they did")

    def test_viewer_can_read_groups_but_not_change_them(self):
        """Decluttering is a reading problem, so every role sees the boxes."""
        self.save()
        self.make_user("viewer3", self.company, role="viewer")
        self.client.logout()
        self.client.login(username="viewer3", password="pw")
        self.assertEqual(self.client.get(self.url("api_chart_groups")).status_code, 200)
        self.assertEqual(self.save(name="Sneaky").status_code, 403)

    def test_branch_admin_cannot_group_outside_their_branch(self):
        self.make_user("branch4", self.company, role="admin", branch="E2")
        self.client.logout()
        self.client.login(username="branch4", password="pw")
        self.assertEqual(self.save().status_code, 200)
        self.assertEqual(
            self.save(parent_employee_id="E1", name="Execs", member_ids=["E5"]).status_code,
            403)


# ===========================================================================
# Phase 3 — the chart page and the client/server metric parity
# ===========================================================================

class ChartPageTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)

    def chart(self, **params):
        url = reverse("org_view:company_detail", kwargs={"slug": self.company.slug})
        if params:
            url += "?" + "&".join(f"{k}={v}" for k, v in params.items())
        return self.client.get(url)

    def test_chart_page_renders_mode_switch_for_admin(self):
        self.make_user("admin3", self.company, role="admin")
        self.client.login(username="admin3", password="pw")
        self.assertContains(self.chart(), "oc-mode-switch")

        for username, role in (("viewer2", "viewer"), ("restricted3", "restricted")):
            self.make_user(username, self.company, role=role)
            self.client.logout()
            self.client.login(username=username, password="pw")
            self.assertNotContains(self.chart(), "oc-mode-switch")

    def test_teams_button_is_reachable_in_every_mode(self):
        """Grouping was previously buried in the side panel, which meant it was
        unreachable in View mode entirely. It's a reading aid — it has to have a
        toolbar entry point wherever an edit role can see the chart."""
        self.make_user("admin8", self.company, role="admin")
        self.client.login(username="admin8", password="pw")
        for mode in ("view", "correct"):
            self.assertContains(self.chart(mode=mode), 'id="oc-groups-btn"', msg_prefix=mode)

        # Read-only roles can see boxes but not build them, so no button.
        self.make_user("viewer4", self.company, role="viewer")
        self.client.logout()
        self.client.login(username="viewer4", password="pw")
        self.assertNotContains(self.chart(), 'id="oc-groups-btn"')

    def test_chart_page_config_has_no_pay_for_restricted(self):
        self.make_user("restricted4", self.company, role="restricted")
        self.client.login(username="restricted4", password="pw")
        self.assertContains(self.chart(), "canSeePay:        false")

    def test_chart_page_quotes_branch_root_id(self):
        """An unquoted CharField badge would be a JS SyntaxError."""
        self.make_user("branch2", self.company, role="admin", branch="E2")
        self.client.login(username="branch2", password="pw")
        self.assertContains(self.chart(), 'branchRootId:     "E2"')

    def test_focus_query_param_accepted(self):
        self.make_user("admin4", self.company, role="admin")
        self.client.login(username="admin4", password="pw")
        resp = self.chart(mode="correct", focus="E3")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'focusId:          "E3"')

    def test_client_metrics_match_server(self):
        """The Django half of the parity test.

        The committed fixture's `expected` block is what tree_builder produced
        when it was generated; metrics.parity.mjs asserts metrics.js reproduces
        the same numbers. This half asserts the server hasn't drifted from it.
        """
        path = (Path(settings.BASE_DIR) / "static" / "org_view" / "js"
                / "__tests__" / "fixture.tree.json")
        fixture = json.loads(path.read_text(encoding="utf-8"))
        tree = strip_aggregation(build_tree_from_rows([dict(r) for r in fixture["rows"]]))

        checked = 0
        stack = [tree]
        while stack:
            node = stack.pop()
            self.assertEqual(node["metrics"], fixture["expected"][node["employee_id"]],
                             node["employee_id"])
            checked += 1
            stack.extend(node["children"])
        self.assertEqual(checked, len(fixture["expected"]))


# ===========================================================================
# Phase 5 — mode specifics, corrections review, the retired scenario page
# ===========================================================================

class ModesAndReviewTests(OrgFixtureMixin, TestCase):

    def setUp(self):
        self.company = self.make_company()
        self.snap = self.make_snapshot(self.company)
        self.admin = self.make_user("admin5", self.company, role="admin")
        self.client.login(username="admin5", password="pw")

    def url(self, name, **kw):
        return reverse(f"org_view:{name}", kwargs={"slug": self.company.slug, **kw})

    def test_correct_mode_summary_counts(self):
        """2 orphans + 1 excluded, reconciling against the census total."""
        Employee.objects.create(
            snapshot=self.snap, employee_id="X2", first_name="Ann", last_name="Lost",
            raw_supervisor_id="8888", job_title="Estimator")
        self.snap.employee_count = self.snap.employees.count()
        self.snap.save(update_fields=["employee_count"])
        c = self.correction(
            self.company, "E4", StructureCorrection.Kind.EXCLUDE,
            {"employee_status": "", "raw_supervisor_id": "E2"}, {})
        C.apply_correction(c, self.snap)

        counts = self.client.get(self.url("api_unattached")).json()["counts"]
        self.assertEqual(counts["total_employees"], 8)
        self.assertEqual(counts["excluded"], 1)
        self.assertEqual(counts["orphans"], 2)
        self.assertEqual(counts["rendered"], 5)

    def test_corrections_review_sorts_conflicts_first(self):
        self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"},
            replay_status=StructureCorrection.ReplayStatus.APPLIED)
        self.correction(
            self.company, "E4", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "GONE"},
            replay_status=StructureCorrection.ReplayStatus.CONFLICT)
        rows = self.client.get(self.url("api_corrections")).json()["corrections"]
        self.assertEqual(rows[0]["employee_id"], "E4")
        self.assertEqual(rows[0]["replay_status"], "conflict")
        # Labels are resolved server-side so the client never maps ids to names.
        self.assertIn("Bo Vice (E2)", rows[1]["before_label"])
        self.assertIn("El Three (E5)", rows[1]["after_label"])

    def test_revert_from_review_restores_value(self):
        c = self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"})
        C.apply_correction(c, self.snap)
        resp = self.client.post(
            reverse("org_view:api_revert_correction",
                    kwargs={"slug": self.company.slug, "pk": c.pk}),
            data="{}", content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        c.refresh_from_db()
        self.assertFalse(c.is_active)
        self.assertEqual(
            Employee.objects.get(snapshot=self.snap, employee_id="E3").raw_supervisor_id, "E2")

    def test_accept_source_deactivates_drifted(self):
        """'Accept source' is a revert — the export now agrees, so retire the row."""
        c = self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"},
            replay_status=StructureCorrection.ReplayStatus.DRIFTED)
        C.apply_correction(c, self.snap)
        self.client.post(
            reverse("org_view:api_revert_correction",
                    kwargs={"slug": self.company.slug, "pk": c.pk}),
            data="{}", content_type="application/json")
        c.refresh_from_db()
        self.assertFalse(c.is_active)

    def test_keep_correction_clears_drift(self):
        c = self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"},
            replay_status=StructureCorrection.ReplayStatus.DRIFTED,
            replay_detail="census now says E1")
        resp = self.client.post(
            reverse("org_view:api_keep_correction",
                    kwargs={"slug": self.company.slug, "pk": c.pk}),
            data="{}", content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        c.refresh_from_db()
        self.assertEqual(c.replay_status, StructureCorrection.ReplayStatus.APPLIED)
        self.assertEqual(c.replay_detail, "")
        self.assertTrue(c.is_active)

    def test_scenario_detail_redirects_to_chart(self):
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Reorg")
        resp = self.client.get(reverse(
            "org_view:scenario_detail",
            kwargs={"slug": self.company.slug, "scenario_id": sc.id}))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp["Location"],
                         f"/org-view/c/{self.company.slug}/?mode=scenario&scenario={sc.id}")

    def test_branch_admin_scenario_tree_is_scoped(self):
        """The privilege gap the old scenario_detail had — regression test.

        api_scenario_tree honoured branch_root_employee_id; the HTML page ignored
        it and rendered the whole company. Now the chart is the only scenario UI
        and gets its tree from the API, so the gap closes by construction.
        """
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Reorg")
        self.make_user("branch3", self.company, role="admin", branch="E2")
        self.client.logout()
        self.client.login(username="branch3", password="pw")
        tree = self.client.get(reverse(
            "org_view:api_scenario_tree",
            kwargs={"slug": self.company.slug, "scenario_id": sc.id})).json()["tree"]
        self.assertEqual(tree["employee_id"], "E2")
        self.assertEqual(_flatten_ids(tree), {"E2", "E3", "E4"})

    def test_scenario_rename_works_from_chart_header(self):
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Old")
        resp = self.client.post(
            reverse("org_view:scenario_action",
                    kwargs={"slug": self.company.slug, "scenario_id": sc.id}),
            {"action": "rename", "name": "FY27 Reorg", "description": "Consolidate ops"})
        self.assertEqual(resp.status_code, 302)
        sc.refresh_from_db()
        self.assertEqual(sc.name, "FY27 Reorg")
        self.assertEqual(sc.description, "Consolidate ops")

    def test_eliminate_records_prior_supervisor(self):
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Reorg")
        S.eliminate_position(sc.positions.get(employee_id="E2"))
        for eid in ("E3", "E4"):
            child = sc.positions.get(employee_id=eid)
            self.assertEqual(child.raw_supervisor_id, "E1")
            self.assertEqual(child.prior_supervisor_id, "E2")

    def test_corrections_review_page_renders(self):
        self.correction(
            self.company, "E3", StructureCorrection.Kind.REPARENT,
            {"raw_supervisor_id": "E2"}, {"raw_supervisor_id": "E5"})
        CorrectionReplayLog.objects.create(
            company=self.company, snapshot=self.snap, applied_count=1)
        resp = self.client.get(self.url("corrections_review"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Latest replay report")
        self.assertContains(resp, "Corrections ledger")

    def test_scenario_action_no_longer_accepts_structural_edits(self):
        """save_position / eliminate / reassign are deleted, not merely unused.

        save_position ran reassign_manager first and let the outer
        `except ValueError` swallow a rejected move, so edit_position never ran
        and the user's attribute edits vanished. That branch no longer exists.
        """
        sc = S.create_scenario(company=self.company, base_snapshot=self.snap, name="Reorg")
        pos = sc.positions.get(employee_id="E3")
        self.client.post(
            reverse("org_view:scenario_action",
                    kwargs={"slug": self.company.slug, "scenario_id": sc.id}),
            {"action": "save_position", "position_id": pos.id, "job_title": "Hacked"})
        pos.refresh_from_db()
        self.assertEqual(pos.job_title, "Technician")
